from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.db import IntegrityError

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.views.generic import UpdateView

from .models import Product, Client
from .forms import ProductForm


def home(request):
    return render(request, 'pages/home.html')


def login_user(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Logged In!')
            return redirect('home')
        else:
            messages.success(request, 'Please try again.....')
            return redirect('login')
    else:
        return render(request, 'pages/login.html')


def logout_user(request):
    logout(request)
    messages.success(request, 'Success Logout!')
    return redirect('home')


def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        client_name = request.POST.get('client_name')
        client_phone = request.POST.get('client_phone')

        if form.is_valid():
            # Check if Client exists
            client, created = Client.objects.get_or_create(
                name=client_name,
                phone=client_phone
            )

            product = form.save(commit=False)
            product.client = client
            product.save()
            messages.success(request, 'Product created successfully!')
            return redirect('product_list')
    else:
        form = ProductForm()

    return render(request, 'pages/product_form.html', {
        'form': form,
        'title': 'Create New Product',
    })


def product_list(request):
    query = request.GET.get('q', '')
    products = Product.objects.filter(
        Q(name__icontains=query) |
        Q(serial_number__icontains=query) |
        Q(client__phone__icontains=query) |
        Q(client__name__icontains=query)  # Add this line to search by client's phone number
    ).select_related('client')  # Use select_related to optimize database queries

    paginator = Paginator(products, 10)  # Show 10 products per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'pages/product_table.html', {'page_obj': page_obj, 'query': ""})


def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            try:
                # Extract client information
                client_name = request.POST.get('client_name')
                client_phone = request.POST.get('client_phone')

                # Try to get an existing client with the same phone number
                client, created = Client.objects.get_or_create(
                    phone=client_phone,
                    defaults={'name': client_name}
                )

                if not created:
                    # If client already exists, update the name
                    client.name = client_name
                    client.save()

                # Save the product and associate with the client
                updated_product = form.save(commit=False)
                updated_product.client = client
                updated_product.save()

                messages.success(request, 'Product updated successfully.')
                return redirect(reverse('product_list'))
            except IntegrityError:
                messages.error(request, 'An error occurred. The phone number might be associated with another client.')
    else:
        form = ProductForm(instance=product)

    context = {
        'form': form,
        'title': 'Edit Product',
        'client_name': product.client.name if product.client else '',
        'client_phone': product.client.phone if product.client else '',
    }
    return render(request, 'pages/product_form.html', context)


# Delete View
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect(reverse('product_list'))
    return render(request, 'pages/product_confirm_delete.html', {'product': product})


def client_list(request):
    query = request.GET.get('q', '')
    clients = Client.objects.annotate(product_count=Count('products')).filter(
        Q(name__icontains=query) | Q(phone__icontains=query)
    ).order_by('name')

    paginator = Paginator(clients, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'pages/client_table.html', {'page_obj': page_obj, 'query': ''})


def client_detail(request, pk):
    client = get_object_or_404(Client, pk=pk)
    products = client.products.all()

    paginator = Paginator(products, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'pages/client_detail.html', {
        'client': client,
        'page_obj': page_obj,
        'title': f"Client Details: {client.name}"
    })


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'pages/product_detail.html', {'product': product})


class ClientUpdateView(UpdateView):
    model = Client
    template_name = 'pages/client_update.html'
    fields = ['name', 'phone']
    success_url = reverse_lazy('client_list')


def exel_page(request):
    return render(request, 'pages/exel/exel_download.html')


# ###################################################################    ##############
# import openpyxl
# from django.http import HttpResponse
# from openpyxl.styles import Font
# from .models import Product
# import random
# from datetime import datetime, timedelta
# from django.shortcuts import render
# from django.utils.timezone import now
# from django.utils.timezone import make_aware
#
#
# def generate_captcha(request):
#     """CAPTCHA misolini generatsiya qiladi va sessiyada saqlaydi."""
#     num1 = random.randint(1, 10)
#     num2 = random.randint(1, 10)
#     request.session['captcha_result'] = num1 + num2
#     return {'num1': num1, 'num2': num2}
#
#
# def validate_captcha(request):
#     """Foydalanuvchi yuborgan CAPTCHA natijasini tekshiradi."""
#     user_result = int(request.POST.get('captcha_result', 0))
#     correct_result = request.session.get('captcha_result')
#     return user_result == correct_result
#
#
# def check_request_limit(request):
#     """So'rovlar limitini tekshiradi va foydalanuvchini bloklaydi."""
#     current_time = now()
#     request_times = request.session.get('request_times', [])
#
#     # So'nggi 1 daqiqadagi so'rovlarni tekshirish
#     request_times = [
#         make_aware(datetime.strptime(time, "%Y-%m-%d %H:%M:%S"))
#         for time in request_times if
#         make_aware(datetime.strptime(time, "%Y-%m-%d %H:%M:%S")) > current_time - timedelta(minutes=1)
#     ]
#
#     # Agar limit oshgan bo'lsa, foydalanuvchini bloklash
#     if len(request_times) >= 3:
#         block_until = current_time + timedelta(hours=1)
#         request.session['blocked_until'] = block_until.strftime("%Y-%m-%d %H:%M:%S")
#         return block_until
#
#     # So'rov vaqtini yangilash
#     request_times.append(current_time.strftime("%Y-%m-%d %H:%M:%S"))
#     request.session['request_times'] = request_times
#     return None
#
#
# def export_to_excel():
#     """Mahsulotlarni Excel faylga eksport qiladi."""
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Products"
#
#     # Sarlavhalarni qo'shish
#     sheet.append([
#         "ID", "Product Name", "Serial Number", "Client Name", "Sold Date",
#         "Warranty Period (Months)", "Is Warranty Active", "Created At"
#     ])
#
#     # Sarlavhalarni qalin qilish
#     for cell in sheet[1]:
#         cell.font = Font(bold=True)
#
#     # Ustun kengliklarini o'rnatish
#     sheet.column_dimensions['A'].width = 10
#     sheet.column_dimensions['B'].width = 20
#     sheet.column_dimensions['C'].width = 20
#     sheet.column_dimensions['D'].width = 25
#     sheet.column_dimensions['E'].width = 15
#     sheet.column_dimensions['F'].width = 25
#     sheet.column_dimensions['G'].width = 20
#     sheet.column_dimensions['H'].width = 20
#
#     # Mahsulot ma'lumotlarini qo'shish
#     for product in Product.objects.all():
#         sheet.append([
#             product.id,
#             product.name,
#             product.serial_number,
#             product.client.name if product.client else "No Client",
#             product.sold_date,
#             product.warranty_period,
#             "Active" if product.is_warranty_active else "Expired",
#             product.created_at.replace(tzinfo=None),
#         ])
#
#     # Javobni tayyorlash
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename=products.xlsx'
#     workbook.save(response)
#     return response
#
#
# def export_products_to_excel(request):
#     blocked_until = request.session.get('blocked_until')
#     if blocked_until:
#         blocked_until = make_aware(datetime.strptime(blocked_until, "%Y-%m-%d %H:%M:%S"))
#         if blocked_until > now():
#             messages.error(request, 'You are temporarily blocked! Please try again later.')
#             return redirect('home')
#
#     if request.method == "GET":
#         captcha = generate_captcha(request)
#         return render(request, 'pages/export_captcha.html', captcha)
#
#     if request.method == "POST":
#         if not validate_captcha(request):
#             captcha = generate_captcha(request)
#             return render(request, 'pages/export_captcha.html', {**captcha, 'error': 'Incorrect CAPTCHA! Try again.'})
#
#         block_until = check_request_limit(request)
#         if block_until:
#             messages.error(request, f'You are temporarily blocked until {block_until}.')
#             return redirect('home')
#
#         # CAPTCHA to'g'ri, Excel faylni yaratish va yuklash
#         response = export_to_excel()
#         messages.success(request, 'Excel file has been successfully generated and downloaded.')
#         return response  # Excel faylni to'g'ridan-to'g'ri yuklash


#  yangisi
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font
from .models import Product
import random
from datetime import datetime, timedelta
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils.timezone import now, make_aware
from django.urls import reverse
import io


def generate_captcha(request):
    num1 = random.randint(1, 10)
    num2 = random.randint(1, 10)
    request.session['captcha_result'] = num1 + num2
    return {'num1': num1, 'num2': num2}


def validate_captcha(request):
    user_result = int(request.POST.get('captcha_result', 0))
    correct_result = request.session.get('captcha_result')
    return user_result == correct_result


def check_request_limit(request):
    current_time = now()
    request_times = request.session.get('request_times', [])
    request_times = [
        make_aware(datetime.strptime(time, "%Y-%m-%d %H:%M:%S"))
        for time in request_times if
        make_aware(datetime.strptime(time, "%Y-%m-%d %H:%M:%S")) > current_time - timedelta(minutes=1)
    ]
    if len(request_times) >= 3:
        block_until = current_time + timedelta(hours=1)
        request.session['blocked_until'] = block_until.strftime("%Y-%m-%d %H:%M:%S")
        return block_until
    request_times.append(current_time.strftime("%Y-%m-%d %H:%M:%S"))
    request.session['request_times'] = request_times
    return None


def generate_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Products"

    headers = [
        "ID", "Product Name", "Serial Number", "Client Name", "Sold Date",
        "Warranty Period (Months)", "Is Warranty Active", "Created At"
    ]
    sheet.append(headers)

    # Sarlavhalarni qalin qilish
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Ustun kengliklarini o'rnatish
    column_widths = [10, 20, 20, 25, 15, 25, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for product in Product.objects.all():
        row = [
            product.id,
            product.name,
            product.serial_number,
            product.client.name if product.client else "No Client",
            product.sold_date,
            product.warranty_period,
            "Active" if product.is_warranty_active else "Expired",
            product.created_at.replace(tzinfo=None),
        ]
        sheet.append(row)

    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    return excel_file


def export_products_to_excel(request):
    blocked_until = request.session.get('blocked_until')
    if blocked_until:
        blocked_until = make_aware(datetime.strptime(blocked_until, "%Y-%m-%d %H:%M:%S"))
        if blocked_until > now():
            messages.error(request, 'You are temporarily blocked! Please try again later.')
            return redirect('home')

    if request.method == "GET":
        captcha = generate_captcha(request)
        return render(request, 'pages/exel/export_captcha.html', captcha)

    if request.method == "POST":
        if not validate_captcha(request):
            captcha = generate_captcha(request)
            return render(request, 'pages/exel/export_captcha.html',
                          {**captcha, 'error': 'Incorrect CAPTCHA! Try again.'})

        block_until = check_request_limit(request)
        if block_until:
            messages.error(request, f'You are temporarily blocked until {block_until}.')
            return redirect('home')

        # Excel faylni yaratish va yuklash
        excel_file = generate_excel()
        response = HttpResponse(excel_file.getvalue(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=products.xlsx'

        # Xabarni qo'shish
        messages.success(request, 'Excel file has been successfully generated and downloaded.')

        # JavaScript orqali fayl yuklanishini boshlab, so'ng home sahifasiga yo'naltirish
        return render(request, 'pages/exel/download_and_redirect.html', {
            'file_url': reverse('download_excel'),
            'redirect_url': reverse('home')
        })


def download_excel(request):
    excel_file = generate_excel()
    response = HttpResponse(excel_file.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    return response



